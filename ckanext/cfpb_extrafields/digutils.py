from openpyxl import load_workbook
try:
    from ckan.plugins.toolkit import Invalid
except ImportError: # pragma: no cover
    # If the custom exception can't be imported, use a more generic exception
    # This happens when ckan is not installed locally, like when running unit tests on travis.
    Invalid = Exception

from ckanext.cfpb_extrafields import validators as v
from ckanext.cfpb_extrafields import options

# Helper functions to convert values of sheets
def strfy(val):
    if isinstance(val, basestring):
        if val.lower().strip() in ["na", "n/a", "not applicable", "select one"]:
            return ""
        return val.strip()
    else:
        return ""

def access_restrictions(ws):
    restrictions = []
    for field, cell in [("Supervisor", "B16"), ("Data Owner", "D16")]:
        if ws[cell].value == "Yes":
            restrictions.append("{0} approval required".format(field))
    addl_restrictions = strfy(ws["B17"].value)
    if addl_restrictions:
        restrictions.append(addl_restrictions)
    return ", ".join(restrictions)

# FIELDS values can be functions that take in a worksheet and provide a value
# These factories return functions to use there.
# For instance, pra_exclusion is in fields D38 and D39.
# Therefore, we want the FIELDS value to be a function that takes in a worksheet and returns
# those cells added together.
# Instead of making this function manually with `dig_func = lambda ws: ws["D38"]+ws["D39"]`,
# we can say `dig_func = concat(["D38", "D39"])` and it will make the function and return it to us.
def concat(fields):
    def concat_fields(ws):
        result = ""
        for field in fields:
            result += strfy(ws[field].value)
        return result
    return concat_fields

def date(cell):
    def get_date(ws):
        val = ws[cell].value
        if hasattr(val, "strftime"):
            val = val.strftime("%Y-%m-%d")
        val = strfy(val) # Treat "n/a" as empty string
        _ = v.reasonable_date_validator(val) # Make sure it's a valid date, but return the string.
        return val
    return get_date

def lower(cell):
    def get_lower(ws):
        return (ws[cell].value or "").lower()
    return get_lower

def normalize(string):
    return (string or "").lower().strip()

def option(cell, opts, default=""):
    """Make sure that a cell's value is one of a set of options.
    This is necessary for cells that are rendered as dropdowns in the Data Catalog
    Since these dropdowns are case-sensitive, we need to confirm that the case of the
    cell value matches the case expected by the dropdown"""
    # Make a map from the normalized options to their actual values
    option_map = dict((normalize(opt), opt) for opt in opts)
    def get_option(ws):
        val = ws[cell].value or ""
        # Normalize the value and look it up in the map to get the proper casing
        return option_map.get(normalize(val), default)
    return get_option

get_transfer_method = option("B48", options.TRANSFER_METHOD, "Other")

# Maps field name to either a cell or a function that's passed the worksheet and should return the value
# Note that some values are currently blank and commented out as they don't map to any fields in the DIG excel sheet
FIELDS = {
    "access_restrictions": access_restrictions,
    "contact_primary_name": "F16",
    # "contact_secondary_name": "B6",
    "data_source_names": "D10",
    # "dataset_notes": "",
    "dig_id": lambda ws: v.dig_id_validator(strfy(ws["B5"].value)),
    "initial_purpose_for_intake": "H15",
    "legal_authority_for_collection": "B25",
    "notes": "H4",
    "pra_exclusion": concat(["D38", "B39"]),
    "pra_omb_control_number": lambda ws: v.pra_control_num_validator(strfy(ws["F37"].value)),
    "pra_omb_expiration_date": date("F38"),
    "privacy_contains_pii": lower("B29"),
    "privacy_has_direct_identifiers": lower("B30"),
    "privacy_has_privacy_act_statement": lower("D30"),
    "privacy_pia_notes": "B33",
    "privacy_pia_title": "D32",
    "privacy_sorn_number": "D31",
    "private": lambda ws: True,# Always have new data sources default to private
    "procurement_document_id": "F24",
    "relevant_governing_documents": "D24",
    "sensitivity_level": "B13",
    "title": "B4",
    "transfer_details": "B54",
    "transfer_initial_size": "B47",
    "transfer_method": get_transfer_method,
    "update_frequency": "F47",
    "usage_restrictions":  concat(["B18", "B19"]),
    # "website_url": "",
    # "wiki_link": "",
}

def get_field(worksheet, field, fields=FIELDS):
    cell_or_func = fields[field]
    if callable(cell_or_func):
        return cell_or_func(worksheet)
    else:
        return strfy(worksheet[cell_or_func].value)

def make_rec_from_sheet(ws, fields):
    result = {}
    errors = []
    for field in fields:
        try:
            result[field] = get_field(ws, field, fields)
        except Invalid as  err:
            errors.append(field + ": " + getattr(err, "error", getattr(err, "message", "UKNOWN_ERROR")))
    return result, errors

def make_rec(excel_file, fields=FIELDS):
    wb = load_workbook(excel_file, read_only=True)
    ws = wb.worksheets[0]
    return make_rec_from_sheet(ws, fields)
